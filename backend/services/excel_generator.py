import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Fixed columns always shown first, in this order
FIXED_COLUMNS = [
    "source_filename",
    "employee_name",
    "employee_id",
    "employer_name",
    "pay_period_month",
    "pay_period_year",
    "currency",
    "base_salary",
    "gross_pay",
    "total_deductions",
    "net_pay",
    "income_tax",
    "national_insurance",
    "health_insurance",
    "pension_employee",
    "pension_employer",
]

COLUMN_LABELS = {
    "source_filename": "File",
    "employee_name": "Employee Name",
    "employee_id": "Employee ID",
    "employer_name": "Employer",
    "pay_period_month": "Month",
    "pay_period_year": "Year",
    "currency": "Currency",
    "base_salary": "Base Salary",
    "gross_pay": "Gross Pay",
    "total_deductions": "Total Deductions",
    "net_pay": "Net Pay",
    "income_tax": "Income Tax",
    "national_insurance": "National Insurance",
    "health_insurance": "Health Insurance",
    "pension_employee": "Pension (Employee)",
    "pension_employer": "Pension (Employer)",
}

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
MONEY_KEYWORDS = {"salary", "pay", "tax", "insurance", "pension", "deduction", "payment", "deductions"}


def is_money_col(col: str) -> bool:
    return any(k in col.lower() for k in MONEY_KEYWORDS)


def generate_excel(rows: list[dict]) -> bytes:
    # Remove internal _raw field
    clean_rows = [{k: v for k, v in row.items() if k != "_raw"} for row in rows]

    # Collect ALL keys across all rows
    all_keys = list(dict.fromkeys(
        FIXED_COLUMNS +
        [k for row in clean_rows for k in row if k not in FIXED_COLUMNS]
    ))

    # Build normalized rows
    normalized = [{col: row.get(col) for col in all_keys} for row in clean_rows]
    df = pd.DataFrame(normalized, columns=all_keys)

    # Rename fixed columns to friendly labels; leave dynamic ones as-is
    rename_map = {k: v for k, v in COLUMN_LABELS.items() if k in df.columns}
    df.rename(columns=rename_map, inplace=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Payslips")
        ws = writer.sheets["Payslips"]

        for col_idx, col_name in enumerate(df.columns, start=1):
            # Header style
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-fit width
            max_len = max(
                len(str(col_name)),
                *[len(str(v)) if v is not None else 0 for v in df[col_name]],
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

            # Money format
            if is_money_col(str(col_name)):
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

    return buffer.getvalue()
